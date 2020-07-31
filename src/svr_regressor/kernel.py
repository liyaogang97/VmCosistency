from sklearn.svm import SVR

import pymysql
from pymysql.cursors import DictCursor
import numpy as np
import openpyxl
import math
from openpyxl.styles import PatternFill

list_regressor_data = []
db = pymysql.connect("localhost", "root", "me521..", "vmconsistency")
cursor = db.cursor(DictCursor)
cursor.execute('select * from kernel')
# 查询后的字段名称可以有cursor.description
# for col in cursor.description:
#     print(col)
kernel_recordings = cursor.fetchall()

for kernel_recording in kernel_recordings:
    regressor_data = {}
    stream_select_sql = "select * from stream where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and type=" + str(kernel_recording['type'])
    cursor.execute(stream_select_sql)
    stream_select_result = cursor.fetchall()

    fio_read_select_sql = "select * from fio_read where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and type=" + str(kernel_recording['type'])
    cursor.execute(fio_read_select_sql)
    fio_read_select_result = cursor.fetchall()

    fio_write_select_sql = "select * from fio_write where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and type=" + str(kernel_recording['type'])
    cursor.execute(fio_write_select_sql)
    fio_write_select_result = cursor.fetchall()

    linpack_select_sql = "select * from linpack where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and type=" + str(kernel_recording['type'])
    cursor.execute(linpack_select_sql)
    linpack_select_result = cursor.fetchall()

    pi5000_select_sql = "select * from pi5000 where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and type=" + str(kernel_recording['type'])
    cursor.execute(pi5000_select_sql)
    pi5000_select_result = cursor.fetchall()

    if (len(stream_select_result) == 1 and len(linpack_select_result) == 1 and len(fio_read_select_result) == 1 and len(
            fio_write_select_result) == 1):
        regressor_data.update(stream_select_result[0])
        regressor_data.update(linpack_select_result[0])
        regressor_data.update(fio_read_select_result[0])
        regressor_data.update(fio_write_select_result[0])
        regressor_data.update(pi5000_select_result[0])
        regressor_data.update(kernel_recording)
        list_regressor_data.append(regressor_data)
        # print(regressor_data)

attributes = ["type", "cpu_number", "memory_count", "triad", "real", "kernel_run_time"]
# attributes=[]
# for key in list_regressor_data[0]:
#     attributes.append(key)
# print(len(attributes))

train_data = []
train_data_target = []
test_data = []
test_data_target = []

for regressor_data in list_regressor_data:
    data = []
    for attribute in attributes:
        data.append(regressor_data[attribute])
    if data[0] == "6230" or data[0] == "8269":
        train_data.append(data)
        train_data_target.append(data[-1])
    else:
        test_data.append(data)
        test_data_target.append(data[-1])

print(len(train_data))
print(len(test_data))

np_train_data = np.array(train_data)
np_test_data = np.array(test_data)

clf = SVR()
clf.fit(np_train_data[:, 1:len(attributes)], train_data_target)
predict_result = clf.predict(np_test_data[:, 1:len(attributes)])

# poly_reg = PolynomialFeatures(degree=2)
# train_poly = poly_reg.fit_transform(np_train_data[:, 1:len(attributes)-1])
# lin_model = linear_model.LinearRegression()
# lin_model.fit(train_poly, train_data_target)
# test_poly = poly_reg.fit_transform(np_test_data[:, 1:len(attributes)-1])
# predict_result = lin_model.predict(test_poly)
# print(predict_result)

row = 1
col = 1
workbook = openpyxl.Workbook()
sheet = workbook.active
for column_name in attributes:
    sheet.cell(row, col, column_name)
    col += 1
sheet.cell(row, col, "预测值")
col += 1
sheet.cell(row, col, "预测误差百分比")
col += 1
sheet.cell(row, col, "预测误差")

for index in range(0, len(test_data)):
    row += 1
    for col in range(0, len(test_data[index])):
        # print(test_data[index])
        sheet.cell(row, col + 1, test_data[index][col])

    col = len(test_data[index]) + 1
    sheet.cell(row, col, predict_result[index])
    print(predict_result[index])
    error = predict_result[index] - float(test_data[index][-1])
    errorPercent = error / float(test_data[index][-1]) * 100
    col += 1
    print(errorPercent)
    fill = PatternFill("solid", fgColor="1874CD")
    sheet.cell(row, col, errorPercent)
    col += 1
    sheet.cell(row, col, error)
    if math.fabs(errorPercent) > 5:
        sheet.cell(row, col).fill = fill
        sheet.cell(row, col - 1).fill = fill

workbook.save("kernel_data.xlsx")
