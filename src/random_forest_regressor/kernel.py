from sklearn.ensemble import RandomForestRegressor
import pymysql
from pymysql.cursors import DictCursor
import numpy as np
import openpyxl
import math
from openpyxl.styles import PatternFill
from sklearn.preprocessing import PolynomialFeatures
from sklearn import metrics


#  MAPE和SMAPE
def mape(y_true, y_pred):
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    y_pred.astype(np.float64)
    y_true.astype(np.float64)
    print(y_pred.dtype)
    print(y_true.dtype)
    return np.mean(np.abs((y_pred - y_true) / y_true)) * 100


def smape(y_true, y_pred):
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    y_pred.astype(np.float)
    y_true.astype(np.float)
    return 2.0 * np.mean(np.abs(y_pred - y_true) / (np.abs(y_pred) + np.abs(y_true))) * 100


list_regressor_data = []
db = pymysql.connect("localhost", "root", "me521..", "vmconsistency")
cursor = db.cursor(DictCursor)
cursor.execute('select * from kernel')
# 查询后的字段名称可以有cursor.description
# for col in cursor.description:
#     print(col)
kernel_recordings = cursor.fetchall()

for kernel_recording in kernel_recordings:
    regressor_data = {}
    stream_select_sql = "select * from stream where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and server_type=" + str(kernel_recording['type'])
    cursor.execute(stream_select_sql)
    stream_select_result = cursor.fetchall()

    fio_read_select_sql = "select * from fio_read where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and type=" + str(kernel_recording['type'])
    cursor.execute(fio_read_select_sql)
    fio_read_select_result = cursor.fetchall()

    fio_write_select_sql = "select * from fio_write where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and type=" + str(kernel_recording['type'])
    cursor.execute(fio_write_select_sql)
    fio_write_select_result = cursor.fetchall()

    linpack_select_sql = "select * from linpack where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and type=" + str(kernel_recording['type'])
    cursor.execute(linpack_select_sql)
    linpack_select_result = cursor.fetchall()

    pi5000_select_sql = "select * from pi5000 where cpu_number=" + str(
        kernel_recording['cpu_number']) + " and cpu_frequency=" + str(
        kernel_recording['cpu_frequency']) + " and memory_count=" + str(
        kernel_recording['memory_count']) + " and type=" + str(kernel_recording['type'])
    cursor.execute(pi5000_select_sql)
    pi5000_select_result = cursor.fetchall()


    if (len(stream_select_result) == 1 and len(fio_read_select_result) == 1 and len(
            fio_write_select_result) == 1 and len(linpack_select_result) == 1):
        regressor_data.update(stream_select_result[0])
        regressor_data.update(fio_read_select_result[0])
        regressor_data.update(fio_write_select_result[0])
        regressor_data.update(kernel_recording)
        regressor_data.update(pi5000_select_result[0])
        regressor_data.update(linpack_select_result[0])
        list_regressor_data.append(regressor_data)

attributes = ["type", "cpu_number", "memory_count", "real", "gflops", "kernel_run_time"]
print(len(attributes))

train_data = []
train_data_target = []
test_data = []
test_data_target = []

for regressor_data in list_regressor_data:
    data = []
    for attribute in attributes:
        data.append(regressor_data[attribute])
    if data[0] == "8260" or data[0] == "8269":
        train_data.append(data)
        train_data_target.append(float(data[-1]))
    else:
        test_data.append(data)
        test_data_target.append(float(data[-1]))

print(len(train_data))
print(len(test_data))

np_train_data = np.array(train_data)
np_test_data = np.array(test_data)

# rfr = RandomForestRegressor(n_estimators=200, oob_score=True)
#
# print(attributes[1:len(attributes) - 1])
# rfr.fit(np_train_data[:, 1:len(attributes) - 1], train_data_target)
# predict_result = rfr.predict(np_test_data[:, 1:len(attributes) - 1])
# print(predict_result)

poly_feature = PolynomialFeatures(degree=4)
poly_train_data = poly_feature.fit_transform(np_train_data[:, 1:len(attributes) - 1])
poly_test_data = poly_feature.fit_transform(np_test_data[:, 1:len(attributes) - 1])
# print(np_train_data[:, 1:len(attributes) - 1])
# lin_reg = LinearRegression()
# lin_reg.fit(poly_train_data, train_data_target)
# predict_result = lin_reg.predict(poly_test_data)
# print(lin_reg.coef_)
rfr = RandomForestRegressor()
rfr.fit(poly_train_data, train_data_target)
predict_result = rfr.predict(poly_test_data)
print("MSE 均方误差:   " + str(metrics.mean_squared_error(test_data_target, predict_result)))
print("MAE  平均绝对误差:  " + str(metrics.mean_absolute_error(test_data_target, predict_result)))
print("R2 决定系数:   " + str(metrics.r2_score(test_data_target, predict_result)))
print("RMSE 均方根误差  " + str(np.sqrt(metrics.mean_squared_error(test_data_target, predict_result))))
print("MAPE 平均绝对百分比误差  " + str(mape(test_data_target, predict_result)))
print("SMAPE 对称平均绝对百分比误差  " + str(smape(test_data_target, predict_result)))

row = 1
col = 1
workbook = openpyxl.Workbook()
sheet = workbook.active
for column_name in attributes:
    sheet.cell(row, col, column_name)
    col += 1
sheet.cell(row, col, "预测值")
col += 1
sheet.cell(row, col, "预测误差百分比")
col += 1
sheet.cell(row, col, "预测误差")

for index in range(0, len(test_data)):
    row += 1
    for col in range(0, len(test_data[index])):
        # print(test_data[index])
        sheet.cell(row, col + 1, test_data[index][col])

    col = len(test_data[index]) + 1
    sheet.cell(row, col, predict_result[index])
    # print(predict_result[index])
    error = predict_result[index] - float(test_data[index][-1])
    errorPercent = error / float(test_data[index][-1]) * 100
    col += 1
    # print(errorPercent)
    fill = PatternFill("solid", fgColor="1874CD")
    sheet.cell(row, col, errorPercent)
    col += 1
    sheet.cell(row, col, error)
    if math.fabs(errorPercent) > 10:
        sheet.cell(row, col).fill = fill
        sheet.cell(row, col - 1).fill = fill

workbook.save("kernel_data.xlsx")
