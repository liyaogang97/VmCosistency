from sklearn.ensemble import RandomForestRegressor
import pymysql
from pymysql.cursors import DictCursor
import numpy as np
import openpyxl
import math
from openpyxl.styles import PatternFill
from sklearn import metrics
from sklearn.model_selection import GridSearchCV
from sklearn.linear_model import LinearRegression
from sklearn.tree import DecisionTreeRegressor
from sklearn.linear_model import Ridge
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import PolynomialFeatures
from sklearn import metrics
from sklearn.feature_selection import SelectKBest, SelectPercentile, f_regression, mutual_info_regression, chi2
import warnings
from sklearn.model_selection import train_test_split
import decimal

warnings.filterwarnings('ignore')


#  MAPE和SMAPE
def mape(y_true, y_pred):
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    y_pred.astype(np.float64)
    y_true.astype(np.float64)
    print(y_pred.dtype)
    print(y_true.dtype)
    return np.mean(np.abs((y_pred - y_true) / y_true)) * 100


def smape(y_true, y_pred):
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    y_pred.astype(np.float)
    y_true.astype(np.float)
    return 2.0 * np.mean(np.abs(y_pred - y_true) / (np.abs(y_pred) + np.abs(y_true))) * 100


train_data = []
train_data_target = []
test_data = []
test_data_target = []
attributes = ["server_type", "cpu_number", "reciprecal_real", "memory_count", "triad"]
print(attributes)

# # 列表中元素的数据类型为字典
# list_regressor_data = []
# db = pymysql.connect("localhost", "root", "me521..", "vmconsistency")
# # 查询返回字典值
# cursor = db.cursor(DictCursor)
# cursor.execute('select * from specjbb2005')
# # 查询后的字段名称可以有cursor.description
# # for col in cursor.description:
# #     print(col)
# specjbb2005_recordings = cursor.fetchall()
#
# for specjbb2005_recording in specjbb2005_recordings:
#     regressor_data = {}
#     # 内存
#     stream_select_sql = "select * from stream where cpu_number=" + str(
#         specjbb2005_recording['cpu_number']) + " and cpu_frequency=" + str(
#         specjbb2005_recording['cpu_frequency']) + " and memory_count=" + str(
#         specjbb2005_recording['memory_count']) + " and server_type=" + str(specjbb2005_recording['server_type'])
#     cursor.execute(stream_select_sql)
#     stream_select_result = cursor.fetchall()
#     # 磁盘I/O读
#     fio_read_select_sql = "select * from fio_read where cpu_number=" + str(
#         specjbb2005_recording['cpu_number']) + " and cpu_frequency=" + str(
#         specjbb2005_recording['cpu_frequency']) + " and memory_count=" + str(
#         specjbb2005_recording['memory_count']) + " and server_type=" + str(specjbb2005_recording['server_type'])
#     cursor.execute(fio_read_select_sql)
#     fio_read_select_result = cursor.fetchall()
#     # 磁盘I/O写
#     fio_write_select_sql = "select * from fio_write where cpu_number=" + str(
#         specjbb2005_recording['cpu_number']) + " and cpu_frequency=" + str(
#         specjbb2005_recording['cpu_frequency']) + " and memory_count=" + str(
#         specjbb2005_recording['memory_count']) + " and server_type=" + str(specjbb2005_recording['server_type'])
#     cursor.execute(fio_write_select_sql)
#     fio_write_select_result = cursor.fetchall()
#
#     linpack_select_sql = "select * from linpack where cpu_number=" + str(
#         specjbb2005_recording['cpu_number']) + " and cpu_frequency=" + str(
#         specjbb2005_recording['cpu_frequency']) + " and memory_count=" + str(
#         specjbb2005_recording['memory_count']) + " and server_type=" + str(specjbb2005_recording['server_type'])
#     cursor.execute(linpack_select_sql)
#     linpack_select_result = cursor.fetchall()
#
#     # CPU计算
#     pi5000_select_sql = "select * from pi5000 where cpu_number=" + str(
#         specjbb2005_recording['cpu_number']) + " and cpu_frequency=" + str(
#         specjbb2005_recording['cpu_frequency']) + " and memory_count=" + str(
#         specjbb2005_recording['memory_count']) + " and server_type=" + str(specjbb2005_recording['server_type'])
#     cursor.execute(pi5000_select_sql)
#     pi5000_select_result = cursor.fetchall()
#
#     kernel_select_sql = "select * from kernel where cpu_number=" + str(
#         specjbb2005_recording['cpu_number']) + " and cpu_frequency=" + str(
#         specjbb2005_recording['cpu_frequency']) + " and memory_count=" + str(
#         specjbb2005_recording['memory_count']) + " and server_type=" + str(specjbb2005_recording['server_type'])
#     cursor.execute(kernel_select_sql)
#     kernel_select_result = cursor.fetchall()
#
#     if (len(stream_select_result) == 1 and len(linpack_select_result) == 1 and len(fio_read_select_result) == 1 and len(
#             fio_write_select_result) == 1 and len(pi5000_select_result) == 1 and len(kernel_select_result) == 1):
#
#         regressor_data.update(stream_select_result[0])
#         regressor_data.update(fio_read_select_result[0])
#         regressor_data.update(fio_write_select_result[0])
#         regressor_data.update(pi5000_select_result[0])
#         regressor_data.update(linpack_select_result[0])
#         regressor_data.update(kernel_select_result[0])
#         regressor_data.update(specjbb2005_recording)
#         list_regressor_data.append(regressor_data)


# np_list_regressor_data = np.array(list_regressor_data)
# np.save("np_list_regressor_data.npy", np_list_regressor_data)
np_list_regressor_data = np.load("np_list_regressor_data.npy")
print(len(np_list_regressor_data))

labels = []
for regressor_data in np_list_regressor_data:
    labels.append(float(regressor_data['thrput']))

train, test, train_data_target, test_data_target = train_test_split(np_list_regressor_data, labels, test_size=0.3)

for iter_train in train:
    data = []
    for attribute in attributes:
        data.append(iter_train[attribute])
    train_data.append(data)


for iter_test in test:
    data = []
    for attribute in attributes:
        data.append(iter_test[attribute])
    # print(data)
    test_data.append(data)

# for regressor_data in np_list_regressor_data:
#     data = []
#     for attribute in attributes:
#         data.append(regressor_data[attribute])
#     if (data[0] == "6230"):
#         train_data.append(data)
#     if (data[0] == "8260"):
#         test_data.append(data)
#     if (regressor_data['server_type'] == "6230"):
#         train_data_target.append(float(regressor_data['thrput']))
#     if (regressor_data['server_type'] == "8260"):
#         test_data_target.append(float(regressor_data['thrput']))


np_train_data = np.array(train_data)
np_test_data = np.array(test_data)
np_test_data_target = np.array(test_data_target)
np_train_data_target = np.array(train_data_target)

np_train_data.astype(np.float)
np_test_data.astype(np.float)
np_test_data_target.astype(np.float)
np_train_data_target.astype(np.float)

# select_k_feature = SelectPercentile(mutual_info_regression, percentile=90)
# new_train_data = select_k_feature.fit(np_train_data, np_train_data_target)
# attribute_index = select_k_feature.get_support(indices=True)
# print(attribute_index)


# attribute_index = [i + 1 for i in attribute_index]
# for index in attribute_index:
#     print(attributes[index])

poly_feature = PolynomialFeatures(degree=1)
# poly_train_data = poly_feature.fit_transform(np_train_data[:, attribute_index])
# poly_test_data = poly_feature.fit_transform(np_test_data[:, attribute_index])

poly_train_data = poly_feature.fit_transform(np_train_data)
poly_test_data = poly_feature.fit_transform(np_test_data)

# # 线性回归
# lin_reg = LinearRegression()
# lin_reg.fit(poly_train_data, train_data_target)
# predict_result = lin_reg.predict(poly_test_data)
# print(lin_reg.coef_)

# 随机森林回归
rfr = RandomForestRegressor()
rfr.fit(poly_train_data, train_data_target)
predict_result = rfr.predict(poly_test_data)


# 决策树回归
# dtr = DecisionTreeRegressor()
# dtr.fit(poly_train_data, train_data_target)
# predict_result = dtr.predict(poly_test_data)

# print("MSE 均方误差:   " + str(metrics.mean_squared_error(test_data_target, predict_result)))
# print("MAE  平均绝对误差:  " + str(metrics.mean_absolute_error(test_data_target, predict_result)))
print("R2 决定系数:   " + str(metrics.r2_score(test_data_target, predict_result)))
print("RMSE 均方根误差  " + str(np.sqrt(metrics.mean_squared_error(test_data_target, predict_result))))
# print("MAPE 平均绝对百分比误差  " + str(mape(test_data_target, predict_result)))
# print("SMAPE 对称平均绝对百分比误差  " + str(smape(test_data_target, predict_result)))

row = 1
col = 1
workbook = openpyxl.Workbook()
sheet = workbook.active
for column_name in attributes:
    sheet.cell(row, col, column_name)
    col += 1

sheet.cell(row, col, "真值")
col += 1
sheet.cell(row, col, "预测值")
col += 1
sheet.cell(row, col, "预测误差百分比")
col += 1
sheet.cell(row, col, "预测误差")

unqualified_number = 0
for index in range(0, len(test_data)):
    row += 1
    for col in range(0, len(test_data[index])):
        # print(test_data[index])
        sheet.cell(row, col + 1, test_data[index][col])
    col = len(test_data[index]) + 1
    sheet.cell(row, col, test_data_target[index])
    col += 1
    sheet.cell(row, col, predict_result[index])
    # print(predict_result[index])
    error = predict_result[index] - float(test_data_target[index])
    errorPercent = error / float(test_data_target[index]) * 100
    col += 1
    # print(errorPercent)
    fill = PatternFill("solid", fgColor="1874CD")
    sheet.cell(row, col, errorPercent)
    col += 1
    sheet.cell(row, col, error)

    if math.fabs(errorPercent) > 10:
        sheet.cell(row, col).fill = fill
        sheet.cell(row, col - 1).fill = fill
        sheet.cell(row, col - 2).fill = fill
        unqualified_number += 1

print(float(unqualified_number) / len(test_data_target))
workbook.save("specjbb2005_data.xlsx")
